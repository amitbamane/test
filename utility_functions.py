try:
    import traceback
    import os
    import pandas as pd
    import numpy as np
    import openpyxl
    import string
    from datetime import datetime
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

except:
    print("SYSTEM IMPORT ERROR !!!!  \n\n")
    traceback.print_exc()
    exit()


color_coding_red = PatternFill(start_color='E84603', end_color='E84603', fill_type='solid')
color_coding_green = PatternFill(start_color='9DEC67', end_color='9DEC67', fill_type='solid')
color_coding_yellow = PatternFill(start_color='EAE93F', end_color='EAE93F', fill_type='solid')
color_coding_blue = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

def find_header_string_position(sheet_name, str_to_find):
    position = [0, 0]
    for i in range(1, sheet_name.max_row + 1):
        for j in range(1, sheet_name.max_column + 1):
            if (str_to_find.strip().upper() in str(sheet_name.cell(i, j).value).strip().upper()):
#                print("Found the str - ", str_to_find, " at : ", i, " , ", j)
                position = [i, j]
                return position

    print("String not Found : ", str_to_find)
    return None

def find_last_col_position(sheet_name, str_to_find):
    last_col_position = find_header_string_position(sheet_name, str_to_find)
    
    if (last_col_position == None) :
        print("String not Found : ", str_to_find)
        return None
    
    last_col_start_pos = last_col_position[1]
        
    for i in range(last_col_start_pos + 1, sheet_name.max_column + 1):
        if (str_to_find.strip().upper() in str(sheet_name.cell(last_col_position[0], i).value).strip().upper()):
            last_col_position[1] = i
            
    return last_col_position


def format_header_cells(file_sheet, cell_position, no_of_rows_to_merge, no_of_columns_to_merge, color_coding,
                        column_width):
    file_sheet.merge_cells(start_row=cell_position[0], start_column=cell_position[1],
                           end_row=cell_position[0] + no_of_rows_to_merge,
                           end_column=cell_position[1] + no_of_columns_to_merge)
    file_sheet.cell(cell_position[0], cell_position[1]).fill = color_coding
    file_sheet.cell(cell_position[0], cell_position[1]).alignment = Alignment(wrap_text=True, vertical='center')
    file_sheet.column_dimensions[get_column_letter(cell_position[1])].width = column_width

    
def load_df_from_excel_file(file_path, sheet_name_to_read, no_of_rows_to_skip, levels_of_index, no_of_data_rows_to_drop):
    
    raw_df = pd.DataFrame()

    # =========== Read file path data provided in config file =======================
    
    raw_df = pd.read_excel(file_path, sheet_name=sheet_name_to_read, skiprows=no_of_rows_to_skip, header=list(range(0,levels_of_index)))
    
    # =========== Merge the headers into one header =======================
    if(levels_of_index > 1) :
        new_header = []
        headers_count_dict = {}
        for col in raw_df.columns:        
            temp_header = ""
            for i in range(0,levels_of_index) :
                if not (col[i].startswith('Unnamed:') or col[i] == '') :
                    temp_header = temp_header + ' ' + col[i]

            if(temp_header not in headers_count_dict) : 
                headers_count_dict.update({temp_header : 0})
            else :
                headers_count_dict.update({temp_header : (headers_count_dict[temp_header] + 1)})
                temp_header = temp_header + ' ' + str(headers_count_dict[temp_header])
                
            new_header.append(temp_header)

        raw_df.columns = pd.MultiIndex.from_tuples(
            list(zip(new_header, raw_df.columns.get_level_values(levels_of_index-1))))
        raw_df.columns = raw_df.columns.droplevel(level = 1)
    
    # =========== drop irrelevant rows after headers from the df and reset the index =======================
    
    raw_df = raw_df.drop(index = list(range(0, no_of_data_rows_to_drop)))
    raw_df = raw_df.reset_index(drop = True)
    raw_df.columns = raw_df.columns.str.strip()
    
    return raw_df


def load_df_from_excel_file_backup(file_path, sheet_name_to_read, no_of_rows_to_skip, levels_of_index, no_of_data_rows_to_drop):
    
    raw_df = pd.DataFrame()

    # =========== Read file path data provided in config file =======================
    raw_df = pd.read_excel(file_path, sheet_name=sheet_name_to_read, skiprows=no_of_rows_to_skip, header=list(range(0,levels_of_index)))
    
    # =========== Merge the headers into one header =======================
    if(levels_of_index > 1) :
        new_header = []
        for col in raw_df.columns:        
            temp_header = ""
            for i in range(0,levels_of_index) :
                if not (col[i].startswith('Unnamed:') or col[i] == '') :
                    temp_header = temp_header + ' ' + col[i]

            new_header.append(temp_header)

        raw_df.columns = pd.MultiIndex.from_tuples(
            list(zip(new_header, raw_df.columns.get_level_values(levels_of_index-1))))
        raw_df.columns = raw_df.columns.droplevel(level = 1)
    
    # =========== drop irrelevant rows after headers from the df and reset the index =======================
    
    raw_df = raw_df.drop(index = list(range(0, no_of_data_rows_to_drop)))
    raw_df = raw_df.reset_index(drop = True)
    
    # ============ Remove all white spaces from raw_df data frame as well as from column headers of all fields ============    
    
#    for column in raw_df.columns:
#        raw_df[column] = raw_df[column].apply(lambda x: str(x).strip())
        
    raw_df.columns = raw_df.columns.str.strip()
    
    return raw_df
    
def add_row_no_in_dataframe(file_name, sheet_name, final_data_frame, column_header_in_excel_sheet, column_header_in_dataframe):

    file_wb = openpyxl.load_workbook(file_name)
    file_sheet = file_wb[sheet_name]

    column_header_in_excel_sheet_pos = find_header_string_position(file_sheet, column_header_in_excel_sheet)

    list_of_sr_no = []
    list_of_row_no = []

    for i in range(column_header_in_excel_sheet_pos[0] + 1, file_sheet.max_row + 1):

        if (file_sheet.cell(i, column_header_in_excel_sheet_pos[1]).value == None):
            continue

        list_of_row_no.append(i)
        list_of_sr_no.append(file_sheet.cell(i, column_header_in_excel_sheet_pos[1]).value)

    row_no_df = pd.DataFrame(list(zip(list_of_sr_no, list_of_row_no)), columns=['Sr. No.', 'Row No in Excel Sheet'])

    for column in row_no_df.columns:
        row_no_df[column] = row_no_df[column].apply(lambda x: str(x).strip())

    for column in final_data_frame.columns:
        final_data_frame[column] = final_data_frame[column].apply(lambda x: str(x).strip())

    final_data_frame = final_data_frame.merge(row_no_df, left_on="Sr. No.", right_on="Sr. No.")

    final_data_frame['Sr. No.'] = final_data_frame['Sr. No.'].apply(lambda x: int(x))
    final_data_frame['Row No in Excel Sheet'] = final_data_frame['Row No in Excel Sheet'].apply(lambda x: int(x))

    return final_data_frame

def write_alerts_compliance_status_in_excel_sheet(file_name, sheet_name, final_data_frame, compliance_headers_dict, last_col_str, no_of_empty_cols_after_last_col):
    
    file_wb = openpyxl.load_workbook(file_name)
    file_sheet = file_wb[sheet_name]

    last_column_loc = find_last_col_position(file_sheet, last_col_str)
    compliance_headers_list = list(compliance_headers_dict.keys())

    for compliance_header_str in compliance_headers_list :
        if compliance_header_str in final_data_frame.columns:
            
            file_sheet.cell(last_column_loc[0],
                            last_column_loc[1] + no_of_empty_cols_after_last_col + 
                            compliance_headers_dict[compliance_header_str]).value = compliance_header_str
            format_header_cells(file_sheet, [last_column_loc[0], 
                                             last_column_loc[1] + no_of_empty_cols_after_last_col + 
                                             compliance_headers_dict[compliance_header_str]], 1, 0, color_coding_blue, 15)
            for index, row in final_data_frame.iterrows():
                if (row[compliance_header_str] not in [None, 'None']) :
                    file_sheet.cell(row['Row No in Excel Sheet'],
                                    last_column_loc[1] + no_of_empty_cols_after_last_col + 
                                    compliance_headers_dict[compliance_header_str]).value = row[compliance_header_str]

    print("Saving workbook after writing compliance status : " + file_name)
    file_wb.save(file_name)
    print("Workbook saved : ")

# -------------Function 1 and 2-----------------

# # xl = pd.ExcelFile('demo.xlsx')
# # xl = pd.read_excel("demo.xlsx")
# xl = openpyxl.load_workbook("demo.xlsx")
# sheet = xl['Sheet1']
# find = find_last_col_position(sheet ,"ABC")


# # format_header_cells(sheet,[1,1],2,0,color_coding_blue,50)
# a= load_df_from_excel_file('demo.xlsx','Sheet1',1,1,2)
# print(a)
# if find is not None:
#     print(find)
# else:
#     print("String not found")



# -------------Function 3-----------------

# import openpyxl
# from openpyxl.styles import PatternFill, Alignment
# from openpyxl.utils import get_column_letter

# # Load the workbook
# # workbook = openpyxl.Workbook()
# xl = openpyxl.load_workbook("demo.xlsx")
# sheet = xl['Sheet1']
# # Assume the sheet name is 'Sheet1'
# # sheet = workbook['Sheet1']

# # Specify formatting parameters
# cell_position = [1, 1]
# no_of_rows_to_merge = 1
# no_of_columns_to_merge = 2
# color_coding = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Use gold color
# column_width = 15

# # Call the function to format header cells
# format_header_cells(sheet, cell_position, no_of_rows_to_merge, no_of_columns_to_merge, color_coding_red, column_width)

# # Save the workbook
# xl.save('formatted_sheet1.xlsx')

# -------------Function 4-----------------

# file_path = 'demo copy.xlsx'
# sheet_name_to_read = 'Sheet1'
# no_of_rows_to_skip = 2
# levels_of_index = 1
# no_of_data_rows_to_drop = 0

# result_df = load_df_from_excel_file(file_path, sheet_name_to_read, no_of_rows_to_skip, levels_of_index,
#                                     no_of_data_rows_to_drop)

# print(result_df)

# print("____________________")

# # -------------Function 5-----------------

# file_path = 'demo copy.xlsx'
# sheet_name_to_read = 'Sheet1'
# no_of_rows_to_skip = 1
# levels_of_index = 1
# no_of_data_rows_to_drop = 0

# result_df = load_df_from_excel_file_backup(file_path, sheet_name_to_read, no_of_rows_to_skip, levels_of_index,
#                                            no_of_data_rows_to_drop)

# print(result_df)


# -------------Function 6-----------------
# file_name = 'demo2.xlsx'
# sheet_name = 'Sheet1'
# column_header_in_excel_sheet = 'Sr. No. in Excel Sheet'
# column_header_in_dataframe = 'Sr. No.'

# data = {
#     'Sr. No.': ['AAA', 'BCD', "GHJ","ASD","WED"]
# }

# your_dataframe = pd.DataFrame(data)
# print(your_dataframe)
# result_df = add_row_no_in_dataframe(file_name, sheet_name, your_dataframe, column_header_in_excel_sheet, column_header_in_dataframe)

# print(result_df)

# -------------Function 7-----------------

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

# Sample data for final_data_frame
data = {
    'Sr. No.': [101, 102, 103, 104],
    'Name': ['John', 'Alice', 'Bob', 'Eve'],
    'Compliance Header 1': ['Yes', 'No', 'Yes', 'No'],
    'Compliance Header 2': ['No', 'Yes', 'No', 'Yes']
}

final_data_frame = pd.DataFrame(data)

# Sample compliance_headers_dict
compliance_headers_dict = {
    'Compliance Header 1': 1,
    'Compliance Header 2': 2
}

# Sample Excel file name and sheet name
file_name = 'demo2.xlsx'
sheet_name = 'Sheet1'

# Sample last column string and number of empty columns after the last column
last_col_str = 'Last Col'
no_of_empty_cols_after_last_col = 2

# Sample color coding for formatting
color_coding_blue = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
write_alerts_compliance_status_in_excel_sheet(file_name, sheet_name, final_data_frame, compliance_headers_dict, last_col_str, no_of_empty_cols_after_last_col)